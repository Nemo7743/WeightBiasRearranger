import openpyxl
import os
from copy import copy

def get_reference_fill(reference_file):
    """
    獲取參考檔案中 'Requant_Sel[3]' 標題儲存格的填滿樣式。
    """
    if not os.path.exists(reference_file):
        return None
    try:
        # 使用 data_only=True 加載以確保獲獲取計算後的樣式 (雖然這對 fill 不一定必要，但在 openpyxl 中較穩健)
        wb = openpyxl.load_workbook(reference_file)
        ws = wb.active
        for r in range(1, 11):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip() == 'Requant_Sel[3]':
                    # 返回該儲存格的樣式複製品
                    return copy(ws.cell(row=r, column=c).fill)
    except Exception as e:
        print(f"Error reading reference color: {e}")
    return None

def modify_excel(file_path, reference_fill=None):
    """
    修改指定的 Excel 檔案，將 'Requant_Sel[3]' 欄位下的 VLIW 指令寫入 '000' 並設定顏色。
    """
    print(f"Loading {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return
        
    ws = wb.active
    
    # 尋找 'Requant_Sel[3]' 標題
    target_col = None
    header_row = None
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() == 'Requant_Sel[3]':
                target_col = c
                header_row = r
                break
        if target_col is not None:
            break
            
    if target_col is None:
        print(f"Error: 找不到 'Requant_Sel[3]' 標題於 {file_path}")
        return
        
    print(f"Found 'Requant_Sel[3]' at Row {header_row}, Column {target_col}")
    
    # 如果有傳入參考顏色則使用，否則使用該檔案自身的標題顏色
    fill_to_use = reference_fill if reference_fill else copy(ws.cell(row=header_row, column=target_col).fill)
    
    modified_count = 0
    start_row = header_row + 1
    
    # 根據 VLIW 5列一組的特性進行修改
    for r in range(start_row, ws.max_row + 1, 5):
        cell = ws.cell(row=r, column=target_col)
        cell.number_format = '@'  # 強制文字格式
        cell.value = '000'
        cell.fill = fill_to_use   # 套用顏色
        modified_count += 1
        
    print(f"Modified {modified_count} rows.")
    
    print(f"Saving {file_path}...")
    try:
        wb.save(file_path)
        print("Done!")
    except Exception as e:
        print(f"Error saving {file_path}: {e}")

if __name__ == "__main__":
    # 獲取腳本所在資料夾路徑，以確保能找到 parent 資料夾下的 Excel 檔案
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 以 Unit0_1_PreP.xlsx 作為顏色的參考來源 (在 parent 資料夾下)
    reference_file = os.path.join(script_dir, "..", "Unit0_1_PreP.xlsx")
    
    ref_fill = get_reference_fill(reference_file)
    if ref_fill:
        print(f"成功從 {reference_file} 取得顏色參考。")
    
    # 僅處理 Unit0_1_PreP.xlsx 檔案
    files_to_process = [reference_file]
    
    if not os.path.exists(reference_file):
        print(f"Error: {reference_file} not found.")
        files_to_process = []

    for f in files_to_process:
        modify_excel(f, reference_fill=ref_fill)
