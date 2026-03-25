import openpyxl
import os
from copy import copy

def get_reference_fill(reference_file):
    """
    獲取參考檔案中 'Requant_Sel[3]' 標題儲存格的填滿樣式。
    """
    if not os.path.exists(reference_file):
        return None
    try:
        # 使用 data_only=True 加載
        wb = openpyxl.load_workbook(reference_file)
        ws = wb.active
        # 在前 11 列尋找標題
        for r in range(1, 12):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None and str(val).strip() == 'Requant_Sel[3]':
                    # 返回該儲存格的樣式複製品
                    return copy(ws.cell(row=r, column=c).fill)
    except Exception as e:
        print(f"Error reading reference color: {e}")
    return None

def modify_conv1_maxpool_excel(file_path, reference_fill=None):
    """
    修改 Unit0_2_Conv1_MaxPool.xlsx 檔案，
    根據 OP_Code (Column C) 的內容修改對應的 Requant_Sel[3] (Column N)。
    """
    print(f"Loading {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return
        
    ws = wb.active
    
    # 1. 尋找 'Requant_Sel[3]' 欄位所在的列 index (Column N = 14)
    target_col = None
    header_row = None
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() == 'Requant_Sel[3]':
                target_col = c
                header_row = r
                break
        if target_col is not None:
            break
            
    if target_col is None:
        print(f"Error: 找不到 'Requant_Sel[3]' 標題於 {file_path}")
        return
        
    print(f"Found 'Requant_Sel[3]' at Row {header_row}, Column {target_col}")
    
    # 顏色參考，優先用傳入的，其次用自身的標題顏色
    fill_to_use = reference_fill if reference_fill else copy(ws.cell(row=header_row, column=target_col).fill)
    
    modified_count = 0
    start_search_row = header_row + 1
    
    # OP_Code 標籤列在第 3 欄 (C)
    tag_col = 3
    
    # 2. 定義標籤對應的數值
    tag_target_val = {
        "input": "000",
        "conv1": "100",
        "max": "000",
        "output": "000"
    }

    print("Scanning for VLIW tags (input/conv1/max/output) in Column C...")
    
    # 遍歷所有列，尋找標籤所在的 OP_Code 列
    for r in range(start_search_row, ws.max_row + 1):
        tag_val = ws.cell(row=r, column=tag_col).value
        if tag_val is None:
            continue
            
        tag_str = str(tag_val).strip().lower()
        
        # 檢查該列是否為目標標籤之一
        for key in tag_target_val:
            if key in tag_str:
                target_val = tag_target_val[key]
                # PreP 資訊在 OP_Code 往後推 4 列 (1:OP_Code, 2:Core, 3:TBO, 4:NoC, 5:PreP)
                prep_row = r + 4
                
                # 安全性檢查 (避免越界或寫到錯的結構，基本上在 Unit0_2 情況下是穩定的)
                cell = ws.cell(row=prep_row, column=target_col)
                cell.number_format = '@' # 文字格式
                cell.value = target_val
                cell.fill = fill_to_use
                
                modified_count += 1
                break
    
    print(f"Total modified Requant_Sel[3] entries: {modified_count}")
    
    print(f"Saving {file_path}...")
    try:
        wb.save(file_path)
        print("Done!")
    except Exception as e:
        print(f"Error saving {file_path}: {e}")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 使用 Unit0_1_PreP.xlsx 作為底色參考
    reference_file = os.path.join(script_dir, "..", "Unit0_1_PreP.xlsx")
    target_file = os.path.join(script_dir, "..", "Unit0_2_Conv1_MaxPool.xlsx")
    
    ref_fill = get_reference_fill(reference_file)
    if ref_fill:
        print(f"成功取得底色參考來自: {os.path.basename(reference_file)}")
    
    if os.path.exists(target_file):
        modify_conv1_maxpool_excel(target_file, ref_fill)
    else:
        print(f"錯誤: 找不到目標檔案 {target_file}")
